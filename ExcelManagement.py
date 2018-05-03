import os
import openpyxl
from pathlib import Path


class WorksheetManager:
    def __init__(self, directory, name):
        if not(os.path.exists(directory)):
            os.makedirs(directory)
        self.path = "{}/{}.xlsx".format(directory, name)
        workbook = Path(self.path)
        if workbook.is_file():
            self.workbook = openpyxl.load_workbook(self.path)
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.save()
        self.sheet = self.workbook.active
        print(self.workbook.sheetnames)

    # Saves any work done
    def save(self):
        self.workbook.save(self.path)

    def create_sheet(self, title):
        self.workbook.create_sheet(title)
        self.save()

    def log_run(self, time_run, ration, complete, amounts):
        row = self.sheet.max_row + 1
        column = openpyxl.utils.column_index_from_string(self.find("Time Run").column)
        self.write_cell(time_run, self.get_cell(column, row))
        self.write_cell(ration, self.get_cell(column + 1, row))
        if complete:
            complete = "Yes"
        else:
            complete = "No"
        self.write_cell(complete, self.get_cell(column + 2, row))
        total = 0
        for key, amount in amounts.items():
            column = openpyxl.utils.column_index_from_string(self.find(key).column)
            self.write_cell(amount.get(), self.get_cell(column, row))
            total += amount.get()
        column = openpyxl.utils.column_index_from_string(self.find("Total").column)
        self.write_cell(total, self.get_cell(column, row))


    # Returns a list of the sheets in the workbook
    def get_sheets(self):
        return self.workbook.get_sheet_names()

    # Returns a sheet object given its name
    def get_sheet(self, name):
        return self.workbook[name]

    # Changes the self.sheet value to the given sheet
    def change_sheet(self, sheet):
        self.sheet = sheet

    # Renames the current sheet to the given value
    def rename_sheet(self, title):
        self.sheet.title = title

    # Returns a cell variable given its column and row as numbers
    def get_cell(self, column, row):
        cell_name = "{}{}".format(openpyxl.utils.get_column_letter(column), str(row))
        # print(cell_name)
        return self.sheet[cell_name]

    # Returns the value in a given cell object
    def read_cell(self, cell):
        return cell.value

    # Writes the given value into the given cell object
    def write_cell(self, value, cell):
        # print(cell)
        cell.value = value

    # Setups sheet i.e. creates layout of file
    def setup_sheet(self, title, headings):
        self.sheet.title = title
        self.write_cell(title, self.sheet["B2"])
        for i, heading in enumerate(headings):
            self.write_cell(heading, self.get_cell(i+2, 4))
            self.write_cell("-", self.get_cell(i+2, 3))
        self.save()

    # Fills a row appropriately with a runs information
    def fill_row(self, ration_name, end_weights, weight_limits, now):
        row = self.sheet.max_row + 1
        self.write_cell(now, self.get_cell(2, row))
        self.write_cell(ration_name, self.get_cell(3, row))
        for i in range(8):
            string = "{}/{}".format(end_weights[i], weight_limits[i])
            self.write_cell(string, self.get_cell(i+4, row))
        self.save()

    # Finds a value in the worksheet and returns its cell
    def find(self, search_string):
        for row in self.sheet.rows:
            for cell in row:
                if self.read_cell(cell) == search_string:
                    return cell
        return None