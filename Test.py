# import excel.excel_management as ex

# ration_logs_ex = ex.WorksheetManager(".", "ration_logs")
# ration_logs_ex.update_sheets("ration_logs")

import openpyxl
import time
from pathlib import Path

def get_cell(sheet, column, row):
  cell_name = "{}{}".format(openpyxl.utils.get_column_letter(column), str(row))
  return sheet[cell_name]

dir = "."
name = "test"
path = "{}/{}.xlsx".format(dir, name)
workbook = Path(path)
if workbook.is_file():
  workbook = openpyxl.load_workbook(path)
else:
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  workbook.save(path)
sheet = workbook.active

cell = get_cell(sheet, 1, 1)
cell.value = "test"
workbook.save(path)

i = 0
while(True):
  print(i)
  i += 1
  time.sleep(1)