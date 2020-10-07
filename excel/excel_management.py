import os
import openpyxl
import itertools
import configparser
import numpy as np
from pathlib import Path

import exceptions as err

class WorksheetManager:
    def __init__(self, directory, name):
        dir = directory
        self.path = "{}/{}.xlsx".format(dir, name)
        workbook = Path(self.path)
        if workbook.is_file():
            self.workbook = openpyxl.load_workbook(self.path)
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.save()
        self.sheet = self.workbook.active

    # Saves any work done
    def save(self):
        self.workbook.save(self.path)

    def create_sheet(self, title):
        self.workbook.create_sheet(title)
        self.save()

    # Returns a list of the sheets in the workbook
    def get_sheets(self):
        return self.workbook.get_sheet_names()

    # Returns a sheet object given its name
    def get_sheet(self, name):
        return self.workbook[name]

    # Changes the self.sheet value to the given sheet
    def change_sheet(self, sheet):
        self.sheet = sheet

    # Renames the current sheet to the given value
    def rename_sheet(self, title):
        self.sheet.title = title

    # Returns a cell variable given its column and row as numbers
    def get_cell(self, column, row):
        cell_name = "{}{}".format(openpyxl.utils.get_column_letter(column), str(row))
        return self.sheet[cell_name]

    # Returns the value in a given cell object
    def read_cell(self, cell):
        return cell.value

    # Writes the given value into the given cell object
    def write_cell(self, value, cell):
        cell.value = value

    # Setups sheet i.e. creates layout of file
    def setup_sheet(self, title, headings):
        self.sheet.title = title
        self.write_cell(title, self.sheet["B2"])
        for i, heading in enumerate(headings):
            self.write_cell(heading, self.get_cell(i+2, 4))
            self.write_cell("-", self.get_cell(i+2, 3))
        self.save()

    # Finds a value in the worksheet and returns its cell
    def find(self, search_string):
        for row in self.sheet.rows:
            for cell in row:
                if self.read_cell(cell) == search_string:
                    return cell
        return None

    def log_run(self, time_run, ration, complete, ingredients, batch_number):
        top_row = self.find("Date Run").row
        row = self.sheet.max_row + 1
        column = self.find("Date Run").column
        self.write_cell(time_run, self.get_cell(column, row))
        self.write_cell(ration, self.get_cell(column + 1, row))
        if complete:
            complete = "Yes"
        else:
            complete = "No"
        self.write_cell(complete, self.get_cell(column + 2, row))
        total = 0
        for ingredient in ingredients:
            try:
                column = self.find(ingredient.name).column
            except AttributeError:
                column = self.find("Total").column
                self.sheet.insert_cols(column)
                self.write_cell("-", self.get_cell(column, top_row - 1))
                self.write_cell(ingredient.name, self.get_cell(column, top_row))
            self.write_cell(ingredient.current_amount, self.get_cell(column, row))
            total += ingredient.current_amount
        column = self.find("Total").column
        self.write_cell(total, self.get_cell(column, row))
        column = self.find("Batch Number").column
        self.write_cell(batch_number, self.get_cell(column, row))

    def update_sheets(self, sheet_type):
        for sheet_name in self.get_sheets():
            sheet = self.get_sheet(sheet_name)
            self.change_sheet(sheet)

            sheet_version = 0
            version_cell = self.find("VERSION")
            if version_cell is not None:
                sheet_version = float(self.read_cell(self.get_cell(version_cell.column + 1, version_cell.row)))
            
            if sheet_version < 1.0:
                sheet_version = 1.0
                version_cell = self._create_version_cell(sheet_version)
                self.save()
            
            if sheet_type == "rations":
                if sheet_version < 2.0:
                    ingredient_cell = self.find("Ingredient")
                    ingredients_table_bounds = self._get_table_bounds(self._get_upper_left_most_cell(ingredient_cell))
                    new_column = ingredient_cell.column + 1
                    for row in self.sheet.rows:
                        previous_cell = row[new_column - 2]
                        cell = row[new_column - 1]
                        if (
                                not (self._cell_is_empty(previous_cell) or self._cell_within_table_bounds(previous_cell, ingredients_table_bounds)) 
                            and not (self._cell_is_empty(cell) or self._cell_within_table_bounds(cell, ingredients_table_bounds))
                            ):
                            origin_cell = self._get_upper_left_most_cell(cell)
                            column_shift = new_column - origin_cell.column
                            self._move_table(origin_cell, column_shift, 0)
                    self.sheet.insert_cols(new_column)
                    self.write_cell("Augar Pin", self.get_cell(new_column, ingredient_cell.row))
                    
                    top_row = ingredient_cell.row
                    column = ingredient_cell.column
                    config = configparser.ConfigParser()
                    config.read("./data/config.ini")
                    if config.has_section("AUGAR_PINS"):
                        for row in itertools.count(top_row + 1):
                            name = self.read_cell(self.get_cell(column, row))
                            if name is None:
                                break
                            augar_pin = config["AUGAR_PINS"].get("{}_pin".format(name.lower().replace(" ", "_")))
                            if augar_pin is not None:
                                self.write_cell(augar_pin, self.get_cell(new_column, row))
                        self.save()

                    sheet_version = 2.0
                    self._update_version_cell(sheet_version, version_cell)
                    self.save()
                
                if sheet_version < 2.1:
                    weigher_cell = self.find("Weighers")
                    weighers_table_bounds = self._get_table_bounds(self._get_upper_left_most_cell(weigher_cell))
                    new_column = weigher_cell.column + 1
                    for row in self.sheet.rows:
                        previous_cell = row[new_column - 2]
                        cell = row[new_column - 1]
                        if (
                                not (self._cell_is_empty(previous_cell) or self._cell_within_table_bounds(previous_cell, weighers_table_bounds)) 
                            and not (self._cell_is_empty(cell) or self._cell_within_table_bounds(cell, weighers_table_bounds))
                            ):
                            origin_cell = self._get_upper_left_most_cell(cell)
                            column_shift = new_column - origin_cell.column
                            self._move_table(origin_cell, column_shift, 0)
                    self.sheet.insert_cols(new_column)
                    self.write_cell("Weigher Pin", self.get_cell(new_column, weigher_cell.row))
                    
                    top_row = weigher_cell.row
                    column = weigher_cell.column
                    config = configparser.ConfigParser()
                    config.read("./data/config.ini")
                    if config.has_section("WEIGHER_PINS"):
                        for row in itertools.count(top_row + 1):
                            weigher_id = self.read_cell(self.get_cell(column, row))
                            if weigher_id is None:
                                break
                            weigher_id = int(weigher_id)
                            weigher_pin = config["WEIGHER_PINS"].get(str(weigher_id))
                            if weigher_pin is not None:
                                self.write_cell(weigher_pin, self.get_cell(new_column, row))
                        self.save()

                    sheet_version = 2.1
                    self._update_version_cell(sheet_version, version_cell)
                    self.save()
                    

            elif sheet_type == "ration_logs":
                pass
    
    def _create_version_cell(self, version):
        num_rows_to_insert = 0
        for row in range(1, 3):
            for column in range(1, 4):
                cell = self.get_cell(column, row)
                if not self._cell_is_empty(cell):
                    num_rows_to_insert = max(num_rows_to_insert, 3 - cell.row)

        self.sheet.insert_rows(1, amount=num_rows_to_insert)
        self.write_cell("VERSION", self.get_cell(1, 1))
        self.write_cell(str(version), self.get_cell(2, 1))
        return self.get_cell(1, 1)

    def _update_version_cell(self, version, version_cell):
        self.write_cell(str(version), self.get_cell(version_cell.column + 1, version_cell.row))

    def _move_table(self, origin_cell, column_shift=0, row_shift=0):
        origin_cell = self._get_upper_left_most_cell(origin_cell)
        if origin_cell.column + column_shift < 1:
            column_shift = 1 - origin_cell.column
        if origin_cell.row + row_shift < 1:
            row_shift = 1 - origin_cell.row
        table_bounds = self._get_table_bounds(origin_cell)
        buffer_zone = self._get_buffer_zone(table_bounds, column_shift, row_shift)

        furthest_column = 0
        for coordinate in buffer_zone:
            furthest_column = max(furthest_column, coordinate[0])

        for coordinate in (buffer_zone):
            cell = self.get_cell(*coordinate)
            if not self._cell_is_empty(cell):
                self._move_table(cell, furthest_column + 1 - cell.column)

        self.sheet.move_range("{}:{}".format(table_bounds[0].coordinate, table_bounds[1].coordinate), rows=row_shift, cols=column_shift)

        self.save()

    def _get_upper_left_most_cell(self, cell):
        if cell.row > 1:
            cell_check = self.get_cell(cell.column, cell.row - 1)
            while not self._cell_is_empty(cell_check):
                cell = cell_check
                if cell.row == 1:
                    break
                cell_check = self.get_cell(cell.column, cell.row - 1)
        if cell.column > 1:
            cell_check = self.get_cell(cell.column - 1, cell.row)
            if not self._cell_is_empty(cell_check):
                while not self._cell_is_empty(cell_check):
                    cell = cell_check
                    if cell.column == 1:
                        break
                    cell_check = self.get_cell(cell.column - 1, cell.row)
                cell = self._get_upper_left_most_cell(cell)
        return cell

    def _get_table_bounds(self, origin_cell):
        origin_column = origin_cell.column
        origin_row = origin_cell.row
        column_bounds = [origin_column, origin_column]
        row_bounds = [origin_row, origin_row]

        all_cells = [origin_cell]
        for cell in all_cells:
            if cell.column > 1:
                cell_check = self.get_cell(cell.column - 1, cell.row)
                if cell_check not in all_cells and not self._cell_is_empty(cell_check):
                    all_cells.append(cell_check)
                    column_bounds[0] = min(cell_check.column, column_bounds[0])
            cell_check = self.get_cell(cell.column + 1, cell.row)
            if cell_check not in all_cells and not self._cell_is_empty(cell_check):
                all_cells.append(cell_check)
                column_bounds[1] = max(cell_check.column, column_bounds[1])
            if cell.row > 1:
                cell_check = self.get_cell(cell.column, cell.row - 1)
                if cell_check not in all_cells and not self._cell_is_empty(cell_check):
                    all_cells.append(cell_check)
                    row_bounds[0] = min(cell_check.row, row_bounds[0])
            cell_check = self.get_cell(cell.column, cell.row + 1)
            if cell_check not in all_cells and not self._cell_is_empty(cell_check):
                all_cells.append(cell_check)
                row_bounds[1] = max(cell_check.row, row_bounds[1])
            
        table_bounds = (self.get_cell(column_bounds[0], row_bounds[0]), self.get_cell(column_bounds[1], row_bounds[1]))
        return table_bounds

    def _get_buffer_zone(self, table_bounds, column_shift=0, row_shift=0):
        row_bounds = [table_bounds[0].row, table_bounds[1].row]
        column_bounds = [table_bounds[1].column, table_bounds[1].column]

        buffer_zone = []
        row_start = row_bounds[0]
        column_start = column_bounds[0]
        if row_start > 1:
            row_start -= 1
        if column_start > 1:
            column_start -= 1

        for row in range(row_start, row_bounds[1] + 2):
            for column in range(column_start, column_bounds[1] + 2):
                buffer_column = column + column_shift
                buffer_row = row + row_shift
                if buffer_column > 0 and buffer_row > 0:
                    if not (buffer_column in range(column_start, column_bounds[1] + 2) and buffer_row in range(row_start, row_bounds[1] + 2)):
                        buffer_zone.append((buffer_column, buffer_row))

        return buffer_zone

    def _cell_within_table_bounds(self, cell, table_bounds):
        return cell.column in range(table_bounds[0].column, table_bounds[1].column + 1) and cell.row in range(table_bounds[0].row, table_bounds[1].row + 1)

    def _cell_is_empty(self, cell):
        return cell.value is None or cell.value == ''
